import re
# импортируем библеотеку, которая обрабатывает регулярные выражения
import openpyxl
#библиотека для импорта в xl


# Импортируем необходимые библиотеки

class Bankmanager():
    # Создаем класс

    def __init__(self):
        self.cardrules = {'720': 'GorgeousBank', '480': 'SuperBank'}
        self.transactionrules = {'Withdrawal': -1, 'Transfer': 1, '+': 1, '-': -1}
        self.menu = '1 - My Current Funds', '2 - Me Expenses For The Period', '3 - Exit The Program'
        self.filepath = 'sms.txt'
        # Описываем свойства объектов в методе

    def del_sms(self, smslist):
        telnum = []
        for i in range(len(smslist)):
            telnum.append(smslist[i].split('//')[1].split(':')[0])
            if telnum[i] not in self.cardrules.keys():
                del smslist[i]
        # Метод используемый для отсеивания сообщений не относящихся к банковским

    def excelexport(self, Received, Spent, Delta):
        book = openpyxl.Workbook()
        sheet = book.active
        Title = sheet.cell(row=1, column=1)
        Title.value = 'Report'
        RecE = sheet.cell(row=2, column=1)
        RecE.value = 'Получено'
        SpentE = sheet.cell(row=3, column=1)
        SpentE.value = 'Потрачено'
        DeltaE = sheet.cell(row=4, column=1)
        DeltaE.value = 'Дельта'
        ResRec = sheet.cell(row=2, column=2)
        ResRec.value = str(Received) + ' EUR'
        ResSpent = sheet.cell(row=3, column=2)
        ResSpent.value = str(Spent) + ' EUR'
        ResDelta = sheet.cell(row=4, column=2)
        ResDelta.value = str(Delta) + ' EUR'
        book.save("Report.xlsx")
        # Метод описывающий экспорт отчета в Excel

    def filereader(self):
        smslist = open(self.filepath)
        smslist = smslist.readlines()
        return smslist
        # Метод считывающий файл

    def datedata(self, smslist):
        date = []
        for i in range(len(smslist)):
            date.append(smslist[i].split('//')[0].split(' ')[0].split('-')[0] + '-'
                        + smslist[i].split('//')[0].split(' ')[0].split('-')[1])
        return (date)
        # Метод выделения из сообщения даты в формате YYYY-MM

    def carddata(self, smslist):
        tel = []
        card = []
        for i in range(len(smslist)):
            tel.append(smslist[i].split('//')[1].split(':')[0])
            tel[i] = self.cardrules.get(tel[i])
            card.append(smslist[i].split('//')[1].split('*')[1][:4])
            card[i] = tel[i] + '*' + card[i]
        return card
        # Метод соединяющий номер карты с номером телефона/названием банка

    def transactionmean(self, smslist):
        transactiontype = []
        for i in range(len(smslist)):
            transactiontype.append(self.transactionrules.get(smslist[i].split('//')[1].split(':')[1])) \
                if smslist[i].split('//')[1].split(':')[0] == '480' \
                else transactiontype.append(self.transactionrules.get(smslist[i].split('//')[1].split(':')[2][1]))
        return transactiontype
        # Метод анализирующий тип транзакции, то есть уменьшение или увеличение баланса

    def balancecheck(self, smslist):
        balance = []
        for i in range(len(smslist)):
            balance.append(smslist[i].split('left:')[1].split(' ')[1]) \
                if smslist[i].split('//')[1].split(':')[0] == '720' \
                else balance.append(smslist[i].split('balance:')[1].split(' ')[1])
        return balance
        # Метод обрабатывающий сообщения для выделения баланса после транзакции

    def operationdelta(self, smslist):
        deltaop = []
        for i in range(len(smslist)):
            deltaop.append(re.findall('(\d+)', smslist[i]))
            deltaop[i] = deltaop[i][8]
        return deltaop
        # Метод определяющий изменение баланса по модулю

    def main(self):
        # Основной метод программы
        SisOn = True
        # Условие работы приложения
        v = self.filereader()
        self.del_sms(v)
        v.sort()
        balance = self.balancecheck(v)
        card = self.carddata(v)
        while SisOn:
            try:
                print(*self.menu, sep="\n")
                choice = int(input('Выберите пункт меню: '))
                # Ввод пункта меню
                if choice is 1:
                    total = 0
                    tmp_balance = {}
                    print('Ваш текущий баланс: ')
                    for i in range(len(v)):
                        tmp_balance.update({card[i]: [balance[i]]})
                    for key, value in tmp_balance.items():
                        print(key + ': ' + str(value[0]))
                        total += int(value[0])
                    print('Всего:', total)
                    anykey = input('Нажмите на любую клавишу ')
                    if anykey is not '':
                        continue
                    # Общий баланс
                if choice is 2:
                    Rec = 0
                    Spent = 0
                    k = 1
                    inp = []
                    date = self.datedata(v)
                    type = self.transactionmean(v)
                    delta = self.operationdelta(v)
                    tmp_date = {}
                    dateinp = input('Введите дату в формате YYYY-MM: ')
                    re.match(r'dddd-dd', dateinp)
                    if re.match(r'\d{4}[-]\d{2}', dateinp):
                    # принт даты по шаблону
                        for i in range(len(v)):
                            if dateinp == date[i]:
                                if not (card[i] in tmp_date.values()):
                                    tmp_date.update({k: card[i]})
                                    inp.append(i)
                                    k += 1
                        for key, value in tmp_date.items():
                            print(key, value)
                        print(k, 'Total')
                        choice2 = int(input('Выберите пункт меню: '))
                        if choice2 in tmp_date.keys():
                            for i in range(len(v)):
                                if dateinp == date[i] and card[i] == tmp_date.get(choice2):
                                    if type[i] is 1:
                                        Rec += int(delta[i])
                                    else:
                                        Spent += int(delta[i])
                            Delta = Rec - Spent
                            print('Получено : ', Rec, 'EUR')
                            print('Потрачено : ', Spent, 'EUR')
                            print('Дельта : ', Delta, 'EUR')
                            while True:
                                exportchoice = input('Вы хотите экспортировать доклад в  Excel?(y/n)')
                                if exportchoice is 'y':
                                    self.excelexport(Rec, Spent, Delta)
                                    break
                                elif exportchoice is 'n':
                                    break
                                else:
                                    print('Ошибка ввода, попробуйте еще раз')
                                    continue
                                # экспорт данных в xl
                        elif choice2 is k:
                            for i in range(len(v)):
                                if dateinp == date[i]:
                                    if type[i] is 1:
                                        Rec += int(delta[i])
                                    else:
                                        Spent += int(delta[i])
                            Delta = Rec - Spent
                            print('Получено : ', Rec)
                            print('Потрачено : ', Spent)
                            print('Дельта : ', Delta)
                        else:
                            print('Такого пункта меню не существует, попробуйте еще раз')
                    else:
                        print('Вы ввели дату не в том формате')
                        continue
                if choice is 3:
                    print('Спасибо за использование нашей программы')
                    SisOn = False
                # Завершение работы автомата
            except:
                print('Такого пункта меню не существует, попробуйте еще раз')
                continue
